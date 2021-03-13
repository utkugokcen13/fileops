import openpyxl
import random
import datetime

def excel_oku():
    wb = openpyxl.load_workbook('D://ogrenci.xlsx')


    sheet = wb["13.03.2021"]

    print(sheet["B2"].value)
    row_count = sheet.max_row
    column_count = sheet.max_column

    for i in range(1,row_count + 1):
        for j in range(1, column_count + 1):
            print(sheet.cell(row=i,column=j).value)






def excel_yaz():

    ad_liste = ["Ali", "Aziz", "Tankut", "Tanju", "Merve", "Nurten", "Türkan"]
    soyad_liste = ["Sancar", "Çolak", "Yüret", "Dilmen", "Terim", "Türkoğlu"]

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet['A1'] = "No"
    sheet['B1'] = "Ad"
    sheet['C1'] = "Soyad"

    # datayı yazıyoruz.
    for i in range(2,5):
        sheet.cell(row=i, column=1).value = i-1
        sheet.cell(row=i, column=2).value = random.choice(ad_liste)
        sheet.cell(row=i, column=3).value = random.choice(soyad_liste)


    # sheet adını değiştir.
    sheet.title = datetime.datetime.now().strftime("%d.%m.%Y")

    # workbook kaydet
    wb.save('D://ogrenci.xlsx')




# Bu çalışan dosyaysa anlamına gelir.
if __name__ == "__main__":
    excel_oku()